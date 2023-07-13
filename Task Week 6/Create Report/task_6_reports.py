from tkinter import *
from tkinter import ttk

from openpyxl import Workbook
wb = Workbook()
ws = wb.active
# ws= wb.save("Report.xlsx")




# from openpyxl import load_workbook
# path = "Report.xlsx"  # existing file to work ( Not a new file. )
# wb = load_workbook(filename=path)
# ws = wb.active  # work on default worksheet




my_w = Tk()
l0 = Label(my_w, text="Cash Flow Report", font=20, width=30, anchor="c")
l0.grid(row=1, column=1, columnspan=4)

l1 = Label(my_w, text="Name: ", width=40, anchor="c")
l1.grid(row=3, column=1)

# add one text box
t1 = Text(my_w, height=1, width=40, bg="white")
t1.grid(row=3, column=2)

l2 = Label(my_w, text="Branch Team member: ", width=40)
l2.grid(row=4, column=1)

# add one text box
t2 = Text(my_w, height=1, width=40, bg="white")
t2.grid(row=4, column=2)


l3 = Label(my_w, text="Department: ", width=40)
l3.grid(row=5, column=1)

# # add one text box
t3 = Text(my_w, height=1, width=40, bg="white")
t3.grid(row=5, column=2)

l4 = Label(my_w, text=" Place: ", width=40)
l4.grid(row=6, column=1)


list_ = ['Chennai Head Office','franchises (Aurangabad)','Chennai','Aurangabad']
place_ = ttk.Combobox(my_w,value=list_,width=50,height=1)
place_.grid(row=6,column=2)

l5 = Label(my_w, text="Fees: ", width=40)
l5.grid(row=7, column=1)

# # add one text box
t5 = Text(my_w, height=1, width=40, bg="white")
t5.grid(row=7, column=2)




b1 = Button(my_w, text="Add Record", width=10, 
	command=lambda: add_data())
b1.grid(row=8, column=2)


def add_data():
    flag_validation = True  # set the flag
    my_name = t1.get("1.0", END)  # read name
    my_branch_team_member = t2.get("1.0", END)  # read class
    my_department = t3.get("1.0", END)
    #   # read mark
    my_place = place_.get()  # read gender
    my_fees = t5.get("1.0", END)
    # length of my_name , my_class and my_gender more than 2
    if len(my_name) < 2 or len(my_branch_team_member) < 2 or len(my_department) < 2:
        flag_validation = False
    try:
        val = int(my_fees)  # checking mark as integer
    except:
        flag_validation = False

    if flag_validation:
        try:
            # i=int(ws.cell(ws.max_row, 1).value or 0)+1
            i=ws.max_row # read the last row number 
            if i!=1:
                i=i+1
            print(str(ws.max_row) + ", i value :"  + str(i))
            l1 = [ my_name, my_branch_team_member, my_fees, my_department,my_place] # row of data to add
            ws.append(l1)
        
            wb.save('Report.xlsx')
            t1.delete("1.0", END)  # reset the text entry box
            t2.delete("1.0", END)  # reset the text entry box
            t3.delete("1.0", END)  # reset the text entry box
            place_.delete("1.0", END)  # reset the text entry box
            t5.delete("1.0", END)  # reset the text entry box
        except Exception as e:
            print(e)

    # else:
    #     l5.grid()
    #     l5.config(fg="red")  # foreground color
    #     l5.config(bg="yellow")  # background color
    #     my_str.set("check inputs.")
    #     l5.after(3000, lambda: l5.config(fg="white", bg="white", text=""))


my_w.mainloop()